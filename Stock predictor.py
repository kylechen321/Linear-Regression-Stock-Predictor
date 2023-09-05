import numpy as np
from sklearn.linear_model import LinearRegression
from matplotlib import pyplot as plt
import openpyxl
from openpyxl import load_workbook
import csv

if __name__ == "__main__":
    
    workbook = load_workbook('Project Data.xlsx')
    worksheet = workbook.active
    end = worksheet.max_row
    xValues = []
    yValues = []

    for row in range(1, end + 1):
        y_value = worksheet.cell(row, 1).value
        if y_value is not None:
            yValues.append(y_value)
            xValues.append(row)
    xValues = np.array(xValues).reshape(-1, 1)
    
    #creating the scatter plot and labels
    plt.scatter(xValues,yValues)
    plt.xlabel('Day')
    plt.ylabel('Stock Price')

    #creating line of best fit with Linear Regression
    model = LinearRegression(fit_intercept = True).fit(xValues, yValues)
    xfit = np.linspace(0, 50)
    yfit = model.predict(xfit[:, np.newaxis])
    plt.plot(xfit,yfit, color = 'r')
    plt.show()

    #Calculating R-Squared
    r_squared = model.score(xValues,yValues)
    print('R-Squared:', r_squared)